from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from datetime import datetime, timedelta, date
from io import BytesIO
import os
import openpyxl
from zoneinfo import ZoneInfo

from backend.dependencies.auth import get_current_user, require_admin
from backend.models.user_model import find_user_by_hash, decrypt_user_for_ui
from backend.models.mobile_ticket_model import get_all_mobile_tickets, update_mobile_ticket_status
from backend.models.vm_ticket_model import get_all_vm_tickets, update_vm_ticket_status
from backend.models.environment_ticket_model import get_all_environment_tickets, update_environment_ticket_status
from backend.utils.email_service import send_email

admin_router = APIRouter()

BRAND = "Fastest | Request HUB"
ALLOWED_STATUSES = {"confirmed", "in_progress", "resolved"}
IST = ZoneInfo("Asia/Kolkata")
UTC = ZoneInfo("UTC")

def _global_admin_email():
    admin_email = os.getenv("ADMIN_EMAIL")
    if not admin_email:
        raise ValueError("ADMIN_EMAIL is not set in backend/.env")
    return admin_email

def _ticket_kind(ticket_id: str):
    if ticket_id.startswith("SR-MD-"):
        return "mobile"
    if ticket_id.startswith("SR-VM-"):
        return "vm"
    if ticket_id.startswith("SR-EN-"):
        return "environment"
    return None

def _user_ui_from_hash(email_hash: str) -> dict | None:
    if not email_hash:
        return None
    u = find_user_by_hash(email_hash)
    if not u:
        return None
    return decrypt_user_for_ui(u)

def _fmt_chennai(dt: datetime | None) -> str:
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(IST).strftime("%d/%m/%Y %H:%M:%S")

def _row(ticket: dict, ticket_type: str) -> dict:
    owner_hash = ticket.get("user_email_hash")
    owner_ui = _user_ui_from_hash(owner_hash)

    raised_by_name = ""
    raised_by_email = ""
    if owner_ui:
        raised_by_name = f"{owner_ui.get('first_name','')} {owner_ui.get('last_name','')}".strip()
        raised_by_email = owner_ui.get("email") or ""

    resolved_by_hash = ticket.get("resolved_by")
    resolved_by_ui = _user_ui_from_hash(resolved_by_hash)
    resolved_by_email = resolved_by_ui.get("email") if resolved_by_ui else None

    return {
        "ticket_id": ticket.get("ticket_id"),
        "type": ticket_type,
        "user_email_hash": owner_hash,
        "raised_by_name": raised_by_name,
        "raised_by_email": raised_by_email,
        "created_at": ticket.get("created_at"),
        "status": ticket.get("status", "confirmed"),
        "resolved_by": resolved_by_hash,
        "resolved_by_email": resolved_by_email,
        "resolved_at": ticket.get("resolved_at"),
    }

@admin_router.get("/tickets", summary="Admin: List all tickets (with decrypted user info)")
def list_all_tickets(user=Depends(get_current_user)):
    require_admin(user)

    data = []
    for t in get_all_mobile_tickets():
        data.append(_row(t, "Mobile"))
    for t in get_all_vm_tickets():
        data.append(_row(t, "Virtual Machine"))
    for t in get_all_environment_tickets():
        data.append(_row(t, "Environment"))

    data.sort(key=lambda x: x.get("created_at") or datetime.min, reverse=True)
    return {"tickets": data}

class StatusUpdateRequest(BaseModel):
    status: str

def _resolved_subject(ticket_id: str) -> str:
    return f"{BRAND} | Ticket Resolved | {ticket_id}"

def _resolved_body(ticket_id: str, ticket_type: str, resolved_by_email: str, hello_name: str) -> str:
    return f"""\
<!doctype html>
<html>
  <body style="font-family: Arial, Helvetica, sans-serif; color:#111827; line-height:1.5;">
    <p>Hello {hello_name},</p>
    <p>Your ticket has been marked as <b>Resolved</b>.</p>

    <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; font-size:14px; width:100%; max-width:650px;">
      <tr>
        <td style="border:1px solid #E5E7EB; padding:8px; width:180px; background:#F9FAFB;"><b>Ticket ID</b></td>
        <td style="border:1px solid #E5E7EB; padding:8px;">{ticket_id}</td>
      </tr>
      <tr>
        <td style="border:1px solid #E5E7EB; padding:8px; width:180px; background:#F9FAFB;"><b>Category</b></td>
        <td style="border:1px solid #E5E7EB; padding:8px;">{ticket_type}</td>
      </tr>
      <tr>
        <td style="border:1px solid #E5E7EB; padding:8px; width:180px; background:#F9FAFB;"><b>Resolved By</b></td>
        <td style="border:1px solid #E5E7EB; padding:8px;">{resolved_by_email}</td>
      </tr>
    </table>

    <p style="margin-top:16px;">Regards,<br/>
    <b>{BRAND}</b></p>
  </body>
</html>
"""

@admin_router.patch("/tickets/{ticket_id}/status", summary="Admin: Update ticket status")
def update_ticket_status(ticket_id: str, req: StatusUpdateRequest, user=Depends(get_current_user)):
    require_admin(user)

    status = (req.status or "").strip()
    if status not in ALLOWED_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status. Use confirmed/in_progress/resolved.")

    kind = _ticket_kind(ticket_id)
    if not kind:
        raise HTTPException(status_code=404, detail="Unknown ticket type for ticket_id.")

    now = datetime.utcnow()

    admin_hash = user.get("sub")
    admin_ui = _user_ui_from_hash(admin_hash)
    admin_email = admin_ui.get("email") if admin_ui else None

    if not admin_hash:
        raise HTTPException(status_code=401, detail="Invalid token (missing sub).")
    if not admin_email:
        raise HTTPException(status_code=404, detail="Admin user profile not found.")

    if kind == "mobile":
        tickets = [t for t in get_all_mobile_tickets() if t.get("ticket_id") == ticket_id]
        if not tickets:
            raise HTTPException(status_code=404, detail="Ticket not found.")
        owner_hash = tickets[0].get("user_email_hash")
        update_mobile_ticket_status(ticket_id, status, resolved_by=admin_hash, resolved_at=now)
        ticket_type = "Mobile"

    elif kind == "vm":
        tickets = [t for t in get_all_vm_tickets() if t.get("ticket_id") == ticket_id]
        if not tickets:
            raise HTTPException(status_code=404, detail="Ticket not found.")
        owner_hash = tickets[0].get("user_email_hash")
        update_vm_ticket_status(ticket_id, status, resolved_by=admin_hash, resolved_at=now)
        ticket_type = "Virtual Machine"

    else:
        tickets = [t for t in get_all_environment_tickets() if t.get("ticket_id") == ticket_id]
        if not tickets:
            raise HTTPException(status_code=404, detail="Ticket not found.")
        owner_hash = tickets[0].get("user_email_hash")
        update_environment_ticket_status(ticket_id, status, resolved_by=admin_hash, resolved_at=now)
        ticket_type = "Environment"

    if status == "resolved" and owner_hash:
        owner_ui = _user_ui_from_hash(owner_hash)
        if owner_ui and owner_ui.get("email"):
            subject = _resolved_subject(ticket_id)
            hello_name = owner_ui.get("first_name") or "there"
            body = _resolved_body(ticket_id, ticket_type, admin_email, hello_name)
            send_email(subject, owner_ui["email"], body, cc_email=_global_admin_email(), is_html=True)

    return {"message": "Status updated", "ticket_id": ticket_id, "status": status}

def _range_from_preset(preset: str) -> tuple[datetime, datetime]:
    """
    Returns [start_utc, end_utc) based on Chennai day boundaries.
    """
    now_ist = datetime.now(IST)

    if preset == "day":
        start_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        end_ist = start_ist + timedelta(days=1)

    elif preset == "week":
        # Monday start
        start_ist = (now_ist - timedelta(days=now_ist.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end_ist = start_ist + timedelta(days=7)

    elif preset == "month":
        start_ist = now_ist.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        # next month
        if start_ist.month == 12:
            end_ist = start_ist.replace(year=start_ist.year + 1, month=1)
        else:
            end_ist = start_ist.replace(month=start_ist.month + 1)

    elif preset == "year":
        start_ist = now_ist.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_ist = start_ist.replace(year=start_ist.year + 1)

    else:
        raise ValueError("Invalid preset")

    return (start_ist.astimezone(UTC), end_ist.astimezone(UTC))

def _range_from_custom(from_str: str, to_str: str) -> tuple[datetime, datetime]:
    """
    from/to are YYYY-MM-DD in Chennai local dates, inclusive.
    Returns [start_utc, end_utc)
    """
    f = date.fromisoformat(from_str)
    t = date.fromisoformat(to_str)
    if t < f:
        raise ValueError("'to' must be >= 'from'")

    start_ist = datetime(f.year, f.month, f.day, 0, 0, 0, tzinfo=IST)
    end_ist = datetime(t.year, t.month, t.day, 0, 0, 0, tzinfo=IST) + timedelta(days=1)
    return (start_ist.astimezone(UTC), end_ist.astimezone(UTC))

@admin_router.get("/tickets/export", summary="Admin: Export tickets to Excel (filtered, Chennai time)")
def export_tickets_excel(
    preset: str = Query(default="month", pattern="^(day|week|month|year|custom)$"),
    from_date: Optional[str] = Query(default=None, alias="from"),
    to_date: Optional[str] = Query(default=None, alias="to"),
    user=Depends(get_current_user),
):
    require_admin(user)

    try:
        if preset == "custom":
            if not from_date or not to_date:
                raise HTTPException(status_code=400, detail="For custom export, 'from' and 'to' are required (YYYY-MM-DD).")
            start_utc, end_utc = _range_from_custom(from_date, to_date)
        else:
            start_utc, end_utc = _range_from_preset(preset)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def in_range(dt: datetime | None) -> bool:
        if not dt:
            return False
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return start_utc <= dt < end_utc

    rows = []
    for t in get_all_mobile_tickets():
        r = _row(t, "Mobile")
        if in_range(r.get("created_at")):
            rows.append(r)
    for t in get_all_vm_tickets():
        r = _row(t, "Virtual Machine")
        if in_range(r.get("created_at")):
            rows.append(r)
    for t in get_all_environment_tickets():
        r = _row(t, "Environment")
        if in_range(r.get("created_at")):
            rows.append(r)

    rows.sort(key=lambda x: x.get("created_at") or datetime.min, reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        "Ticket ID",
        "Type",
        "Raised By (Name)",
        "Raised By (Email)",
        "Created At ",
        "Status",
        "Resolved By (Email)",
        "Resolved At ",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("ticket_id"),
            r.get("type"),
            r.get("raised_by_name"),
            r.get("raised_by_email"),
            _fmt_chennai(r.get("created_at")),
            r.get("status"),
            r.get("resolved_by_email") or "",
            _fmt_chennai(r.get("resolved_at")),
        ])

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 26

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"tickets_{preset}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )